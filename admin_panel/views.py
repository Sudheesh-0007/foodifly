from django.shortcuts import render, redirect
from django.contrib import auth, messages
from django.shortcuts import render, get_object_or_404, redirect
from accounts.models import Account
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib import messages
from django.views.decorators.cache import never_cache
from .decorators import admin_required
from admin_panel.services.report_service import *
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.utils import timezone
from io import BytesIO


@never_cache
def admin_login(request):
    if request.user.is_authenticated:
        if request.user.is_admin:
            return redirect("admin_user_management")
        else:
            messages.warning(request, "You are already logged in as a user.")
            return redirect("home")

    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")
        user = auth.authenticate(email=email, password=password)

        if user is not None and user.is_admin:
            auth.login(request, user)
            return redirect("admin_user_management")
        else:
            messages.error(request, "Admin access only.")
            return redirect("admin_login")

    return render(request, "admin_panel/admin_login.html")


@never_cache
@admin_required
def admin_user_management(request):

    users_list = Account.objects.all().order_by("-date_joined")

    query = request.GET.get("keyword")
    if query:
        users_list = users_list.filter(
            Q(first_name__icontains=query)
            | Q(last_name__icontains=query)
            | Q(email__icontains=query)
        )

    paginator = Paginator(users_list, 10)
    page = request.GET.get("page")

    paged_users = paginator.get_page(page)

    context = {
        "users": paged_users,
        "keyword": query,
        "total_count": Account.objects.count(),
        "active_count": Account.objects.filter(is_active=True).count(),
    }
    return render(request, "admin_panel/user_manage.html", context)


def toggle_user_status(request, user_id):

    user = get_object_or_404(Account, id=user_id)

    if user.is_active:
        user.is_active = False
        messages.success(request, f"User {user.email} has been blocked.")
    else:
        user.is_active = True
        messages.success(request, f"User {user.email} has been unblocked.")

    user.save()
    return redirect("admin_user_management")


@never_cache
def admin_logout(request):
    auth.logout(request)
    messages.success(
        request, "You have been logged out of the Sovereign Management Portal."
    )
    return redirect("admin_login")


def report_dashboard(request):

    filter_type = request.GET.get("filter", "monthly")

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    orders = get_filtered_orders(filter_type, start_date, end_date).select_related(
        "user"
    )

    summary = get_sales_summary(orders)

    recent_orders = get_recent_orders(orders)

    best_products = get_best_selling_products(orders)

    best_categories = get_best_selling_categories(orders)

    chart_data = get_sales_chart_data(filter_type, start_date, end_date)

    context = {
        **summary,
        "recent_orders": recent_orders,
        "best_products": best_products,
        "best_categories": best_categories,
        "filter_type": filter_type,
        "start_date": start_date,
        "end_date": end_date,
        **chart_data,
    }

    return render(request, "admin_panel/admin_reports.html", context)


def export_sales_excel(request):

    filter_type = request.GET.get("filter", "monthly")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    orders = get_filtered_orders(filter_type, start_date, end_date).select_related(
        "user"
    )

    total_orders = orders.count()
    total_sales = orders.aggregate(total=Sum("grand_total"))["total"] or 0
    average_order = total_sales / total_orders if total_orders > 0 else 0

    workbook = openpyxl.Workbook()

    sheet = workbook.active
    sheet.title = "Sales Report"
    sheet["A3"] = "Filter Type"
    sheet["B3"] = filter_type.capitalize()
    sheet["A1"] = "FOODIFLY SALES REPORT"
    sheet.merge_cells("A1:F1")
    sheet["A1"].font = Font(bold=True, size=16)
    sheet["A4"] = "Start Date"
    sheet["B4"] = start_date or "-"

    sheet["A5"] = "End Date"
    sheet["B5"] = end_date or "-"

    sheet["A7"] = "ORDER SUMMARY"

    sheet["A8"] = "Total Sales"
    sheet["B8"] = float(total_sales)

    sheet["A9"] = "Total Orders"
    sheet["B9"] = total_orders

    sheet["A10"] = "Average Order Value"
    sheet["B10"] = float(average_order)

    sheet["A11"] = "Generated On"
    sheet["B11"] = timezone.now().strftime("%Y-%m-%d %H:%M")
    sheet["A12"] = "Generated By"
    sheet["B12"] = request.user.email
    sheet["A1"].alignment = Alignment(horizontal="center")
    headers = [
        "Order Number",
        "Customer",
        "Date",
        "Payment Method",
        "Grand Total",
        "Status",
    ]

    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=14, column=col_num).value = header
    for cell in sheet[14]:

        cell.font = Font(bold=True)
    green_fill = PatternFill(
        start_color="0A2D21", end_color="0A2D21", fill_type="solid"
    )

    for cell in sheet[7]:
        cell.fill = green_fill
        cell.font = Font(bold=True, color="FFFFFF")

    row_num = 15

    for order in orders:

        sheet.cell(row=row_num, column=1).value = order.order_number

        sheet.cell(row=row_num, column=2).value = (
            f"{order.user.first_name} {order.user.last_name}"
        )

        sheet.cell(row=row_num, column=3).value = order.created_at.strftime("%Y-%m-%d")

        sheet.cell(row=row_num, column=4).value = order.payment_method

        sheet.cell(row=row_num, column=5).value = float(order.grand_total)

        sheet.cell(row=row_num, column=6).value = order.status

        row_num += 1

        for col in sheet.columns:

            max_length = 0

            column = col[0].column

            column_letter = get_column_letter(column)

            for cell in col:

                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            sheet.column_dimensions[column_letter].width = max_length + 4

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="sales_report.xlsx"'

    workbook.save(response)

    return response


def export_sales_pdf(request):

    response = HttpResponse(content_type="application/pdf")

    response["Content-Disposition"] = 'attachment; filename="sales_report.pdf"'

    pdf_buffer = BytesIO()

    doc = SimpleDocTemplate(pdf_buffer)

    styles = getSampleStyleSheet()

    elements = []
    filter_type = request.GET.get("filter", "monthly")

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    orders = get_filtered_orders(filter_type, start_date, end_date)

    summary = get_sales_summary(orders)

    elements.append(Paragraph("FOODIFLY SALES REPORT", styles["Title"]))
    elements.append(Paragraph("Premium Food E-Commerce Analytics", styles["Italic"]))
    elements.append(
        Paragraph(
            f"Generated On: {timezone.now().strftime('%d-%m-%Y %H:%M')}",
            styles["Normal"],
        )
    )

    elements.append(Spacer(1, 10))

    elements.append(Paragraph(f"Filter Type: {filter_type.title()}", styles["Normal"]))

    elements.append(Paragraph(f"Start Date: {start_date or '-'}", styles["Normal"]))

    elements.append(Paragraph(f"End Date: {end_date or '-'}", styles["Normal"]))

    elements.append(Spacer(1, 20))

    elements.append(Spacer(1, 20))

    elements.append(Paragraph("Order Summary", styles["Heading2"]))

    elements.append(
        Paragraph(f"Total Sales: ₹{summary['total_sales']}", styles["Normal"])
    )

    elements.append(
        Paragraph(f"Total Orders: {summary['total_orders']}", styles["Normal"])
    )

    elements.append(
        Paragraph(f"Total Users: {summary['total_users']}", styles["Normal"])
    )

    elements.append(
        Paragraph(f"Total Products: {summary['total_products']}", styles["Normal"])
    )

    elements.append(
        Paragraph(f"Total Categories: {summary['total_categories']}", styles["Normal"])
    )

    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Order Details", styles["Heading2"]))

    elements.append(Spacer(1, 10))
    table_data = [
        [
            "Order No",
            "Customer",
            "Date",
            "Payment",
            "Total",
            "Status",
        ]
    ]

    for order in orders.select_related("user"):

        table_data.append(
            [
                order.order_number,
                f"{order.user.first_name} {order.user.last_name}",
                order.created_at.strftime("%d-%m-%Y"),
                order.payment_method,
                f"₹{order.grand_total}",
                order.status,
            ]
        )
    table = Table(table_data)
    table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#0A2D21"),
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white,
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    1,
                    colors.black,
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold",
                ),
                (
                    "BACKGROUND",
                    (0, 1),
                    (-1, -1),
                    colors.whitesmoke,
                ),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)

    pdf = pdf_buffer.getvalue()

    pdf_buffer.close()

    response.write(pdf)

    return response
